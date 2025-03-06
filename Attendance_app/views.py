import uuid
from locations.models import Location
import re
import jdatetime
from asgiref.sync import async_to_sync
from channels.layers import get_channel_layer
from django.shortcuts import render, get_object_or_404, redirect, HttpResponse
from django.urls import reverse
from datetime import datetime, timedelta
from geopy.distance import geodesic
from django.views.decorators.csrf import csrf_exempt
from django.views.generic import TemplateView
from django.views.generic.list import ListView
from jalali_date import datetime2jalali
from django.utils.decorators import method_decorator
from .forms import StaffCreateUser, ShiftWorkForm, PositionForm, ProfileForm, HolidayForm, VacationForm, \
    UpdateProfileForm
from .models import AttendanceUser
from .mixin import CustomizedRquirementLogin
from django.contrib.auth.decorators import login_required, user_passes_test
from pricing.models import Profile, User, CustomUser, Income, ShiftWork, Positions, Holidays, Vacation, \
    VacationType, Day, NoneInProgress
from django.http import JsonResponse, HttpResponse, HttpResponseNotAllowed
from django.utils import timezone
# exel

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .utils import get_jalali_date, get_day_mapping, get_current_shift, handle_non_progress_users, get_month_names, \
    handle_progress_and_none_progress_user, extract_time_ranges
from django.http import HttpResponse
from django.db.models import Q

from .decorators import check_progress, profile_required, subscription_required
from django.db.models.functions import ExtractMonth, ExtractYear
from django.db.models import F
from django.shortcuts import render
from django.utils import timezone
from django.db.models import Q
from django.core.cache import cache
from django.contrib.auth.decorators import login_required
import re

@subscription_required
@profile_required
def restricted_view(request, *args, **kwargs):
    return redirect('Attendance:home')


@check_progress #home page
def create_attendance_view(request):
    """ View for handling attendance and user progress. """

    position = Profile.objects.select_related('user').get(user=request.user)
    now = timezone.now()
    month, year = get_jalali_date()

    # Get user's income (Optimized)
    income = Income.objects.filter(month=month, year=year, user=request.user).values_list('user_income',
                                                                                          flat=True).first()

    if request.user.is_staff:
        location_exists = Location.objects.filter(created_by=request.user).exists()
        users = CustomUser.objects.filter(created_who=request.user)

        # Use caching to prevent redundant processing
        cache_key = f"progress_data_{request.user.id}"
        progress_data = cache.get(cache_key)

        if not progress_data:
            progress_data = handle_progress_and_none_progress_user(users)
            cache.set(cache_key, progress_data, timeout=300)  # Cache for 5 minutes

        return render(request, 'Attendance_app/index.html', {
            'position': position,
            'income': income,
            'month': month,
            'year': year,
            'no_confirmation_users': progress_data['attendance_obj'],
            'in_progress_users': progress_data['in_progress_users'],
            'none_progression_count': progress_data['staff_none_progress_users'],
            'users': users,
            'location': location_exists,
        })

    return render(request, 'Attendance_app/index.html', {
        'position': position,
        'income': income,
        'year': year,
        'month': month,
    })


@method_decorator(check_progress, name='dispatch')
class AttendanceListView(CustomizedRquirementLogin, ListView):
    model = AttendanceUser

    def get_income_or_create(self, attendance_obj):
        """
        Retrieve or create an Income object for the given attendance object.
        """
        profile = attendance_obj.user.possit
        job_time_hours = attendance_obj.job_time.total_seconds() / 3600

        income, created = Income.objects.get_or_create(
            month=attendance_obj.created_date.month,
            year=attendance_obj.created_date.year,
            user=attendance_obj.user,
            defaults={
                "created_date": attendance_obj.created_date,
                "position": profile,
                "job_time": attendance_obj.job_time,
                "user_income": profile.profile_position.position_income * job_time_hours,
                "created_by": self.request.user.created_who
            }
        )
        return income

    def format_job_time(self, job_time_str):
        """
        Format the job time string into a localized format.
        """
        return job_time_str.replace("day", "روز").split(",")[0]

    def calculate_income(self, attendance, position_income):
        """
        Calculate the income for a specific attendance entry.
        """
        return round(position_income * (attendance.job_time.total_seconds() / 3600), 4)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        pk = self.kwargs['pk']
        month = self.kwargs['month']
        year = self.kwargs['year']

        context['months'] = get_month_names()
        user = get_object_or_404(CustomUser, id=pk)

        attendances = AttendanceUser.objects.filter(month=month, year=year, user=user).order_by('-created_date')

        context['user'] = user
        if attendances.exists():
            attendance_obj = attendances.first()
            income = self.get_income_or_create(attendance_obj)
            position_income = income.position.profile_position.position_income

            results = [self.calculate_income(attendance, position_income) for attendance in attendances]
            job_time_list = [self.format_job_time(str(attendance.job_time)) for attendance in attendances]

            context["object"] = zip(attendances, results, job_time_list)
            context['income'] = income
            context['month_job_time'] = self.format_job_time(str(income.job_time))
        else:
            context["object"] = None
            context['income'] = None

        context['month'] = month
        context['year'] = year
        return context


def result_detail(request, pk):
    attendance_obj = AttendanceUser.objects.get(id=pk)

    income, created = Income.objects.get_or_create(
        month=attendance_obj.created_date.month,
        year=attendance_obj.created_date.year,
        user=attendance_obj.user,
        defaults={
            "created_date": attendance_obj.created_date,
            "position": Profile.objects.get(user=attendance_obj.user),
            "job_time": attendance_obj.job_time,
            "user_income": Profile.objects.get(user=attendance_obj.user).profile_position.position_income *
                           (attendance_obj.job_time.total_seconds() / 3600),
            "created_by": request.user.created_who
        }
    )

    zipped_times = extract_time_ranges(attendance_obj.last_info)

    return render(request, 'Attendance_app/result.html', {'zipped_times': zipped_times, 'income': income})


def start_attendance_view(request):

    now = datetime.now()
    date, start = now.date(), now.time()

    # دریافت اطلاعات روز هفته
    day_mapping = get_day_mapping()
    reversed_day_number = day_mapping[now.weekday()]
    current_day = ['شنبه', 'یک‌شنبه', 'دو‌شنبه', 'سه‌شنبه', 'چهارشنبه', 'پنج‌شنبه', 'جمعه'][reversed_day_number]

    # بررسی تعطیلات و موقعیت مکانی کاربر
    check_holidays = Holidays.objects.filter(date=date).exists()
    location = Location.objects.filter(created_by=request.user.created_who, active=True).exists()

    # اطلاعات شیفت کاری
    profile = Profile.objects.get(user=request.user)
    shiftwork = profile.profile_position.shift_work
    current_shift = shiftwork.filter(work_days__day_of_week=reversed_day_number).last()

    work_holi_day = check_holidays or current_shift is None
    overtime = work_holi_day or not (current_shift.work_start_time < start < current_shift.work_end_time) if current_shift else True

    # ارسال شناسه کاربر به WebSocket
    async_to_sync(get_channel_layer().group_send)(
        'group_name', {'type': 'send_user_id', 'user_id': request.user.id}
    )

    # دریافت یا ایجاد رکورد حضور
    attendance_obj, created = AttendanceUser.objects.get_or_create(
        user=request.user, created_date=date,
        defaults={'token': uuid.uuid4(), 'start': start, 'in_progress': True}
    )

    # تنظیم وضعیت حضور
    if created or not attendance_obj.in_progress:
        if not created and attendance_obj.confirmation is None and location:
            return redirect(reverse('Attendance:get_user_location'))

        if not created and not str(attendance_obj.end)[:8] in attendance_obj.last_info:
            attendance_obj.last_info += f"\n start={str(attendance_obj.start)[:8]}, end={str(attendance_obj.end)[:8]}*"

        attendance_obj.start = start
        attendance_obj.in_progress = True

    # تنظیم وضعیت تعطیلی و اضافه‌کاری
    attendance_obj.holiday_check = check_holidays

    if current_shift and attendance_obj.job_time >= current_shift.required_time:
        attendance_obj.overtime_check = overtime
    else:
        attendance_obj.overtime_check = False
    attendance_obj.save(required_time=current_shift.required_time if current_shift else timedelta())

    if location and attendance_obj.confirmation is None:
        return redirect(reverse('locations:get_user_location'))

    return render(request, 'Attendance_app/start.html', {
        'started': attendance_obj.start, 'pk': attendance_obj.token, 'at': attendance_obj,
        'work_in_holi': work_holi_day, 'current_day': current_day, 'monthly': profile.profile_position.monthly
    })


def process_result_view(request, pk):
    attend = get_object_or_404(AttendanceUser, user=request.user, token=pk)
    # To ensure
    if not attend.in_progress:
        month, year = get_jalali_date()
        return redirect(
            reverse("Attendance:result_list", kwargs={"pk": request.user.id, 'month': month, 'year': year}, ))

    attend.in_progress = False
    attend.end = datetime.now().time()
    # Calculate the duration of working time
    job_time = datetime.combine(datetime.min, attend.end) - datetime.combine(datetime.min, attend.start)
    # Calculate the number of hours elapsed since the record was created
    date_hours = datetime.now().date() - attend.created_date

    # Add 24 hours for each day that has passed
    job_time += timedelta(days=date_hours.days)

    attend.job_time += job_time
    attend.save()
    return redirect(reverse("price:procces_pricing", kwargs={'pk': pk}))


@method_decorator(check_progress, name='dispatch')
class ShowResult(TemplateView):
    template_name = 'Attendance_app/result.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        request = self.request

        pk = self.kwargs.get('pk')
        user_id = self.kwargs.get('user', request.user.id)

        token = request.session.get('token')
        attend = get_object_or_404(AttendanceUser, user_id=user_id, id=pk) if pk else get_object_or_404(AttendanceUser,
                                                                                                        user=request.user,
                                                                                                        token=token)

        month, year = get_jalali_date()
        context.update({
            'month': month,
            'year': year,
            'object': attend,
            'user': request.user,
            'date': attend.created_date
        })

        job_time = str(attend.job_time)
        context['job_time'] = job_time.replace("day", "روز")[:14] if 'day' in job_time else job_time[:7]

        context.update({
            'start': attend.start.strftime("%H:%M:%S %p"),
            'end': attend.end.strftime("%H:%M:%S %p"),
        })

        matches = re.findall(r"start=(\d{2}:\d{2}:\d{2}), end=(\d{2}:\d{2}:\d{2})", attend.last_info)
        context['zipped_times'] = matches if matches else []

        income, _ = Income.objects.get_or_create(
            month=attend.created_date.month,
            year=attend.created_date.year,
            user=attend.user,
            defaults={'position': attend.user.possit, 'job_time': timedelta()}
        )

        job_time_duration = datetime.combine(datetime.min, attend.end) - datetime.combine(datetime.min, attend.start)
        context.update({
            'monthly': income.position.profile_position.monthly,
            'income': income,
            'inc': income.user_income if income.position.profile_position.monthly else round(
                income.position.profile_position.position_income * (job_time_duration.total_seconds() / 3600), 4)
        })

        return context

def list_holidays(request):
    staff = request.user
    if not staff.is_staff:
        holidays = Holidays.objects.filter(created_by=staff.created_who)
    else:
        holidays = Holidays.objects.filter(created_by=staff)

    return render(request, 'Attendance_app/holiday_list.html', {'holidays': holidays})


def download_excel_user(request, pk, month, year):
    user = CustomUser.objects.get(id=pk)
    positions = Profile.objects.get(user=user)
    attendances = AttendanceUser.objects.filter(
        month=month,
        year=year,
        user=user
    )
    income = Income.objects.filter(
        month=month,
        year=year,
        user=user
    )
    if not income:
        raise ValueError('no income data')
    MONTH_NAMES = get_month_names()

    # Create an Excel workbook and sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Set month name
    month_name = MONTH_NAMES.get(month)
    sheet["A1"] = f"Month: {month_name}"
    sheet.merge_cells("A1:F1")
    sheet["A1"].font = Font(bold=True)

    # Set table headers
    headers = ["First Name", "Last Name", "Position", "Total Working Time", "Price", "Total Price"]
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = sheet[f"{col_letter}2"]
        cell.value = header
        cell.font = Font(bold=True)
    job_time = []
    objects = attendances

    for obj in objects:
        job = str(obj.job_time)
        if 'day' in job:
            job_time.append(job.replace("day", "روز")[:14])
        else:
            job_time.append(job[:7])
    # Write table data
    row_num = 3
    objects = zip(attendances, job_time)
    for attendance, job_tim in objects:
        sheet[f"A{row_num}"] = user.username
        sheet[f"B{row_num}"] = user.last_name
        sheet[f"C{row_num}"] = positions.profile_position.positions  # Extract the position attribute
        sheet[f"D{row_num}"] = job_tim
        inc = income.position.profile_position.position_income * (attendance.job_time.total_seconds() / 3600)
        sheet[f"E{row_num}"] = round(inc, 4)
        sheet[f"F{row_num}"] = income.user_income
        row_num += 1

    # Set response content type
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    # Provide a filename for the Excel file
    response["Content-Disposition"] = f"attachment; filename={user.username}_info.xlsx"

    # Save the workbook to the response
    workbook.save(response)
    return response


def getting_vacation(request):
    #todo i have to get the limit from manager
    user = request.user
    if request.method == 'POST':
        form = VacationForm(data=request.POST)
        if form.is_valid():
            vacation = form.save(commit=False)
            vacation.user = request.user
            vacation.save()
            userprofile = Profile.objects.get(user=user)
            userprofile.vacation.add(vacation)
            if userprofile.vacation.count() >= 30:
                raise AttendanceUser('تعداد مرخصی های شما بیش از حد مجاز میباشد')
            userprofile.save()

            form.save_m2m()
            return redirect(reverse('Attendance:redirected_view'))
    else:
        form = VacationForm()
    return render(request, 'Attendance_app/get_vacation.html', {'form': form})#todo #



def personal_info(request):
    user = request.user
    profile = Profile.objects.get(user=user)
    position = profile.profile_position
    income = Income.objects.filter(user=user).last()
    return render(request, 'Attendance_app/personal_info.html',
                  {'profile': profile, 'position': position, 'income': income})
