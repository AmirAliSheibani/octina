import uuid
from locations.models import Location
import re
import jdatetime
from asgiref.sync import async_to_sync
from channels.layers import get_channel_layer
from django.shortcuts import render, get_object_or_404, redirect, HttpResponse
from django.urls import reverse
from datetime import datetime, timedelta
from geopy.distance import geodesic
from django.views.decorators.csrf import csrf_exempt
from django.views.generic import TemplateView
from django.views.generic.list import ListView
from jalali_date import datetime2jalali

from Attendance_app.forms import StaffCreateUser, ShiftWorkForm, PositionForm, ProfileForm, HolidayForm, VacationForm, \
    UpdateProfileForm
from Attendance_app.models import AttendanceUser, AttendanceStatus
from Attendance_app.mixin import CustomizedRquirementLogin
from django.contrib.auth.decorators import login_required, user_passes_test
from pricing.models import Profile, User, CustomUser, Income, ShiftWork, Positions, Holidays, Vacation, \
    VacationType, Day, NoneInProgress
from django.http import JsonResponse, HttpResponse, HttpResponseNotAllowed
from django.utils import timezone
# exel

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from Attendance_app.utils import get_jalali_date, get_day_mapping, get_current_shift, handle_non_progress_users, \
    get_month_names, \
    handle_progress_and_none_progress_user, format_jalali_date_with_weekday
from django.http import HttpResponse
from django.db.models import Q

from Attendance_app.decorators import check_progress, profile_required, subscription_required
from django.db.models.functions import ExtractMonth, ExtractYear
from django.db.models import F
from pricing.services.positions import deactivate_position


# Create your views here.


def download_excel(request, pk, month, year):
    if not request.user.is_superuser:
        users = CustomUser.objects.filter(created_who=request.user)
        # positions = Profile.objects.filter(created_by=request.user)
        attendances = AttendanceUser.objects.filter(
            month=month,
            year=year,
            user__in=users
        )
        income = Income.objects.get(
            month=month,
            year=year,
            user__in=users
        )
    else:
        users = CustomUser.objects.filter(created_who__isnull=False).order_by('username')
        # positions = Profile.objects.all()
        attendances = AttendanceUser.objects.filter(
            month=month,
            year=year,
            user__in=users
        )
        income = Income.objects.filter(
            month=month,
            year=year,
            user__in=users
        )

    no_income_users = users.exclude(id__in=income.values_list('user_id', flat=True))
    profile_position = getattr(getattr(users, 'possit', None), 'profile_position', None)
    positions = str(profile_position) if profile_position else ''

    # Create an Excel workbook and sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    MONTH_NAMES = get_month_names()

    # Set month name
    month_name = MONTH_NAMES.get(month)
    sheet["A1"] = f"Month: {month_name}"
    sheet.merge_cells("A1:E1")
    sheet["A1"].font = Font(bold=True)

    # Set table headers
    headers = ["First Name", "Last Name", "Position", "Total Working Time", "Total Price"]
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = sheet[f"{col_letter}2"]
        cell.value = header
        cell.font = Font(bold=True)
    job_time = []
    # مرتب‌سازی لیست attendances بر اساس نام کاربر
    # attendances = sorted(attendances, key=lambda a: a.user.username)

    for obj in attendances:
        job = str(obj.job_time)
        if 'day' in job:
            job_time.append(job.replace("day", "روز")[:14])
        else:
            job_time.append(job[:7])
    # Write table data
    row_num = 3

    # users = sorted(users, key=lambda u: u.username)
    # # مرتب‌سازی لیست income بر اساس نام کاربر
    # income = sorted(income, key=lambda i: i.user.username)
    # Write table data
    row_num = 3
    data = []
    for income_entry in income:
        user = income_entry.user

        data.append((user, income_entry))

    for (user, user_income), w in zip(data, job_time):
        sheet[f"A{row_num}"] = user.username
        sheet[f"B{row_num}"] = user.last_name
        profile_position = getattr(user, 'possit', None).profile_position if hasattr(user, 'possit') else None
        sheet[f"C{row_num}"] = str(profile_position) if profile_position else ''
        sheet[f"D{row_num}"] = str(w)
        sheet[f"E{row_num}"] = str(user_income.user_income)
        row_num += 1

    # Apply styles
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    for user in no_income_users:
        sheet[f"A{row_num}"] = user.username
        sheet[f"B{row_num}"] = user.last_name
        profile_position = getattr(user, 'possit', None).profile_position if hasattr(user, 'possit') else None
        sheet[f"C{row_num}"] = str(profile_position) if profile_position else ''
        sheet[f"A{row_num}"].fill = red_fill
        sheet[f"B{row_num}"].fill = red_fill
        sheet[f"C{row_num}"].fill = red_fill
        row_num += 1

    # Set response content type
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    # Provide a filename for the Excel file
    response["Content-Disposition"] = f"attachment; filename={request.user.username}_users.xlsx"

    # Save the workbook to the response
    workbook.save(response)

    return response


def no_confirmation_check(request):
    """
    View for staff users to see attendance records without confirmation.
    """
    users = CustomUser.objects.filter(created_who=request.user)
    no_confirmation = AttendanceUser.objects.filter(
        user__in=users,
        confirmation__in=[False, None]
    )
    return render(request, 'Attendance_app/confirmation_user.html', {'no_confirmation': no_confirmation})


def accept_confirmation(request, pk):
    """
    View to accept and confirm attendance for a user.
    """
    attendance_obj = get_object_or_404(AttendanceUser, pk=pk)
    attendance_obj.confirmation = True
    attendance_obj.save()
    return redirect('managing:no_confirmation_check')


def not_accepted_confirmation(request, pk):
    """
    View to reject attendance for a user.
    """
    attendance_obj = get_object_or_404(AttendanceUser, pk=pk)
    attendance_obj.delete()
    return redirect('managing:no_confirmation_check')


def absent_record_list_view(request, month, year):
    MONTH_NAMES = get_month_names()
    users = CustomUser.objects.filter(created_who=request.user)

    # فیلتر رکوردهای اون ماه
    records = AttendanceStatus.objects.filter(user__in=users, month=month, year=year)
    unique_days = records.order_by('created_date').values_list('created_date', flat=True).distinct()

    return render(request, 'Attendance_app/absent_users_list.html', {
        'unique_days': unique_days,
        'months': MONTH_NAMES,
        'records_count' : records.count(),
        'month': month,
        'year': year,
    })


def absent_record_detail_view(request,date):
    # users = CustomUser.objects.filter(created_who=request.user)
    Absences_record = AttendanceStatus.objects.filter(created_date=date, user__created_who=request.user)
    # absent_users = Absences_record.user.filter(created_who=request.user)
    month, year = get_jalali_date()
    return render(request, 'Attendance_app/absent_users_detail.html',
                  {'absent_users': Absences_record, 'month': month,
        'year': year,})


def manager_view_absences_for_user(request, user_id, month, year):
    user = get_object_or_404(CustomUser, id=user_id, created_who=request.user)
    MONTH_NAMES = get_month_names()

    absence_records = AttendanceStatus.objects.filter(
        user=user, month=month, year=year
    ).order_by('created_date')
    print(absence_records)

    for record in absence_records:
        record.jalali_str = format_jalali_date_with_weekday(record.created_date)

    return render(request, 'Attendance_app/manager_absent_detail.html', {
        'target_user': user,
        'absence_records': absence_records,
        'months': MONTH_NAMES,
        'month': month,
        'year': year,
    })


def non_progress(request, month, year):
    MONTH_NAMES = get_month_names()
    now = timezone.now()
    users = CustomUser.objects.filter(created_who=request.user)
    filter_non_progress = NoneInProgress.objects.filter(user__in=users, month=month, year=year).exclude(
        created_date=jdatetime.date.fromgregorian(date=now.date()))

    return render(request, 'Attendance_app/non_progress.html',
                  {'filter_non_progress': filter_non_progress, 'months': MONTH_NAMES, 'month': month, 'year': year})


def delete_non_progress(request, pk, month, year):
    non = NoneInProgress.objects.get(id=pk)
    non.delete()
    return redirect(reverse('managing:non_progress', kwargs={'month': month, 'year': year}))


def delete_monthly_non_progress(request, month, year):
    users = CustomUser.objects.filter(created_who=request.user)
    monthly_non = NoneInProgress.objects.filter(user__in=users, month=month, year=year)
    monthly_non.delete()
    return redirect(reverse('managing:non_progress', kwargs={'month': month, 'year': year}))


def in_progress_users(request, month, year):
    if request.user.is_staff:
        # non_progress_records = NoneInProgress.objects.filter(user__in=users, month=month, year=year)
        # non_progress_users = CustomUser.objects.filter(nonprogres__in=non_progress_records).distinct()
        non_progress_users = CustomUser.objects.filter(created_who=request.user, absent=True)
        # in_progress_records = AttendanceUser.objects.filter(user__in=users, absent=False) # todo : i have to ensure user is not absent in models
        in_progress_users = CustomUser.objects.filter(created_who=request.user, absent=False).distinct()

        return render(request, 'Attendance_app/in_progress_users.html',
                      {'in_progress_users': in_progress_users, 'non_progress_users': non_progress_users,
                       'months': get_month_names(), 'month': month, 'year': year})
    return redirect(reverse('Attendance:redirected_view'))


def list_shift_work(request):
    staff = request.user
    shiftwork = ShiftWork.objects.filter(created_by=staff)

    return render(request, 'Attendance_app/shiftwork_list.html', {'shiftwork': shiftwork})


def create_shift_work(request):
    check_Shift = ShiftWork.objects.filter(created_by=request.user).count()
    if check_Shift >= 10:
        check_Shift = True
        return render(request, 'Attendance_app/create_shift.html', {'check_Shift': check_Shift})
    if request.method == 'POST':
        form = ShiftWorkForm(request.POST)
        if form.is_valid():
            shift_work = form.save(commit=False)
            shift_work.created_by = request.user
            shift_work.save()

            shift_work.work_days.set(form.cleaned_data['work_days'])
            return redirect(reverse('managing:list_shift_work'))


    else:
        form = ShiftWorkForm()

    all_days = Day.objects.all()  # Retrieve all Day options

    return render(request, 'Attendance_app/create_shift.html', {'form': form, 'all_days': all_days})


# def check_deleting_shift_work(request, pk):
#     shift = ShiftWork.objects.get(id=pk)


def delete_shift_work(request, pk):
    shift = get_object_or_404(ShiftWork, id=pk)
    if request.method == 'POST':
        related_positions = list(shift.Shift_work.all())
        shift.delete()
        for position in related_positions:
            if position.shift_work.count() == 0:
                deactivate_position(position, 'بدون شیفت کاری')
        return redirect('managing:list_shift_work')

    context = {
        'object_type': 'شیفت کاری',
        'object_name': shift.name,
        'cancel_url': reverse('managing:list_shift_work')
    }
    return render(request, 'Attendance_app/confirm_delete.html', context)



def update_shift_work(request, pk):
    shift = ShiftWork.objects.get(id=pk)

    all_days = Day.objects.all()  # Retrieve all Day options
    selected_days = shift.work_days.all()  # Retrieve currently selected Day options for the shift

    if request.method == 'POST':
        form = ShiftWorkForm(request.POST, instance=shift)
        if form.is_valid():
            print("Form is valid")  # Add a print statement or log message for debugging
            form.save()
            return redirect(reverse('managing:list_shift_work'))
        else:
            print("Form is not valid")
    else:
        form = ShiftWorkForm(instance=shift, initial={'work_days': selected_days})

    return render(request, 'Attendance_app/create_shift.html',
                  {'shift': shift, 'form': form, 'all_days': all_days, 'selected_days': selected_days})


def list_position(request):
    staff = request.user
    position = Positions.objects.filter(created_by=staff)

    return render(request, 'Attendance_app/position_list.html', {'position': position})


def create_position(request):
    check_Positions = Positions.objects.filter(created_by=request.user).count()
    if check_Positions >= 10:
        check_Positions = True
        return render(request, 'Attendance_app/c_position.html', {'check_Positions': check_Positions})
    if request.method == 'POST':
        form = PositionForm(request.POST, request=request)
        if form.is_valid():
            position = form.save(commit=False)
            position.created_by = request.user
            position.save()
            form.save_m2m()  # Save many-to-many relationships

            return redirect(
                reverse('managing:setting_app'))
    else:
        form = PositionForm(request=request)

    return render(request, 'Attendance_app/c_position.html', {'form': form})


def delete_position(request, pk):
    position = Positions.objects.get(id=pk)
    if request.method == 'POST':
        position.delete()
        return redirect(reverse('managing:list_position'))
    context = {
        "object_type": "سمت",
        "object_name": position,
        "cancel_url": reverse('managing:list_position')
    }
    return render(request, 'Attendance_app/confirm_delete.html', context)


def update_position(request, pk):
    position = Positions.objects.get(id=pk)

    if request.method == 'POST':
        form = PositionForm(request.POST, instance=position, request=request)
        if form.is_valid():
            form.save()
            return redirect(reverse('managing:list_position'))
    else:
        form = PositionForm(instance=position, request=request)

    return render(request, 'Attendance_app/c_position.html', {'form': form, 'position': position})


def create_holiday(request):
    jalali_join = datetime2jalali(request.user.date_joined).strftime('%y/%m/%d _ %H:%M:%S')
    # check_profile = Profile.objects.filter(created_by=request.user).count()
    # if check_profile >= 10:
    #     check_profile = True
    #     return render(request, 'Attendance_app/c_profile.html', {'check_profile': check_profile})
    if request.method == 'POST':
        form = HolidayForm(request.POST)
        if form.is_valid():
            shift_work = form.save(commit=False)
            shift_work.created_by = request.user
            shift_work.save()
            return redirect(
                reverse('managing:setting_app'))
    else:
        form = HolidayForm()

    return render(request, 'Attendance_app/c_holiday.html', {'form': form})


def delete_holiday(request, pk):
    holiday = Holidays.objects.get(id=pk)
    if request.method == 'POST':

        holiday.delete()
        return redirect(reverse('Attendance:list_holidays'))
    context = {
        'object_type': "روز تعطیل",
        'object_name': holiday.name,
        'cancel_url': reverse('Attendance:list_holidays')
    }
    return render(request, 'Attendance_app/confirm_delete.html', context)




def update_holiday(request, pk):
    holiday = Holidays.objects.get(id=pk)

    if request.method == 'POST':
        form = HolidayForm(request.POST, instance=holiday)
        if form.is_valid():
            form.save()
            return redirect(reverse('managing:list_holidays'))
    else:
        form = HolidayForm(instance=holiday)

    return render(request, 'Attendance_app/c_holiday.html', {'form': form, 'holiday': holiday})


def staff_user_list(request, pk, month, year):
    MONTH_NAMES = get_month_names()

    if not request.user.is_superuser:
        users = CustomUser.objects.filter(created_who=request.user)
        # positions = Profile.objects.filter(created_by=request.user)
        income = Income.objects.filter(
            month=month,
            year=year,
            user__in=users
        ).order_by('user_income')
        no_income_users = users.exclude(id__in=income.values_list('user_id', flat=True))
        attendances = AttendanceUser.objects.filter(
            month=month,
            year=year,
            user__in=users
        )
    else:
        users = CustomUser.objects.filter(created_who__isnull=False).order_by('username')
        # positions = Profile.objects.all()
        income = Income.objects.filter(user__in=users, month=month, year=year).order_by('user_income')
        no_income_users = users.exclude(id__in=income.values_list('user_id', flat=True))
        attendances = AttendanceUser.objects.filter(
            month=month,
            year=year,
            user__in=users
        )

    job_time = []
    objects = attendances
    if users is None:
        return HttpResponse('شما هیچ یوزری برای نمایش ندارید!')
    if income is None:
        return HttpResponse('یوزر های شما هیچ درآمدی نداشتند!')

    if objects is None:
        return HttpResponse('یوزر های شما هیچ ورود و خروجی نداشتند!')

    for obj in objects:
        job = str(obj.job_time)
        if 'day' in job:
            job_time.append(job.replace("day", "روز")[:14])
        else:
            job_time.append(job[:7])
    data = []
    for income_entry in income:
        user = income_entry.user
        attendance = attendances.filter(user=user).first()

        data.append((user, attendance, income_entry))
    last_name_filter = request.GET.get('last_name_filter')
    job_time_filter = request.GET.get('job_time_filter')
    income_filter = request.GET.get('income_filter')

    if last_name_filter:
        data = [entry for entry in data if entry[0].last_name.startswith(last_name_filter)]

    if job_time_filter:
        if job_time_filter == 'smaller':
            data = sorted(data, key=lambda entry: entry[3].job_time)
        elif job_time_filter == 'greater':
            data = sorted(data, key=lambda entry: entry[3].job_time, reverse=True)

    if income_filter:
        if income_filter == 'smaller':
            data = sorted(data, key=lambda entry: entry[3].user_income)
        elif income_filter == 'greater':
            data = sorted(data, key=lambda entry: entry[3].user_income, reverse=True)

    return render(request, 'Attendance_app/AdminUserList.html',
                  {'object': data, 'checkincome': no_income_users,
                   'noIncome': no_income_users,
                   'months': MONTH_NAMES, 'month': month, 'year': year})


def create_user_for_staff(request):
    check_User = CustomUser.objects.filter(created_who=request.user).count()
    if check_User >= 10:
        check_User = True
        return render(request, 'Attendance_app/create_user.html', {'check_User': check_User})

    staff = request.user
    if request.method == 'POST':
        form = StaffCreateUser(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.created_who = staff
            user.set_password(form.cleaned_data['password2'])
            user.save()

            return redirect(reverse('managing:setting_app'))

    else:
        form = StaffCreateUser()

    return render(request, 'Attendance_app/create_user.html', {'form': form})


def delete_user_for_staff(request, pk, mo, year):
    user = CustomUser.objects.get(id=pk)
    if request.method == 'POST':

        user.delete()
        return redirect(reverse('managing:user_list', kwargs={'pk': pk, 'month': mo, 'year': year}))

    context = {
        "object_type": "کاربر",
        "object_name": user.username,
        "cancel_url": reverse('managing:user_list', kwargs={'pk': pk, 'month': mo, 'year': year})
    }
    return render(request, 'Attendance_app/confirm_delete.html', context)


def update_user_for_staff(request, pk):
    user = CustomUser.objects.get(id=pk)

    profile = Profile.objects.get(user=user)
    if request.method == 'POST':
        form = StaffCreateUser(request.POST, instance=user)
        if form.is_valid():
            user.username = form.cleaned_data.get('username')
            user.email = form.cleaned_data.get('email')
            user.first_name = form.cleaned_data.get('first_name')
            user.last_name = form.cleaned_data.get('last_name')
            user.set_password(form.cleaned_data.get('password2'))
            user.save()
            month, year = get_jalali_date()
            return redirect(reverse('managing:user_list', kwargs={'pk': pk, 'month': month, 'year': year}))
        form2 = UpdateProfileForm(request.POST, instance=profile)
        if form2.is_valid():
            form2.save()
            return redirect(reverse('managing:list_profile'))

    else:
        form = StaffCreateUser(instance=user)
        form2 = UpdateProfileForm(instance=profile, request=request)

    return render(request, 'Attendance_app/create_user.html', {'form': form, 'form2': form2, 'user': user})


def list_profile(request):
    staff = request.user
    profile = Profile.objects.filter(created_by=staff)

    return render(request, 'Attendance_app/profile_list.html', {'profile': profile})


def create_profile(request):
    check_profile = Profile.objects.filter(created_by=request.user).count()
    if check_profile >= 10:
        check_profile = True
        return render(request, 'Attendance_app/c_profile.html', {'check_profile': check_profile})
    if request.method == 'POST':
        form = ProfileForm(request.POST, request=request)
        if form.is_valid():
            shift_work = form.save(commit=False)
            shift_work.created_by = request.user
            shift_work.save()
            return redirect(
                reverse('managing:setting_app'))
    else:
        form = ProfileForm(request=request)

    return render(request, 'Attendance_app/c_profile.html', {'form': form})


def update_profile(request, pk):
    profile = Profile.objects.get(id=pk)

    if request.method == 'POST':
        form = UpdateProfileForm(request.POST, instance=profile)
        if form.is_valid():
            form.save()
            return redirect(reverse('managing:list_profile'))
    else:
        form = UpdateProfileForm(instance=profile, request=request)

    return render(request, 'Attendance_app/c_profile.html', {'form': form, 'profile': profile})


def delete_profile(request, pk):
    profile = Profile.objects.get(id=pk)
    if request.method == 'POST':

        profile.delete()
        return redirect(reverse('managing:list_profile'))

    context = {
        "object_type": "پروفایل",
        "object_name": profile,
        "cancel_url": reverse('managing:list_profile')
    }
    return render(request, "Attendance_app/confirm_delete.html", context)


def setting_app(request):
    users = CustomUser.objects.filter(created_who=request.user)

    profile_users = users.filter(possit__isnull=True)

    month, year = get_jalali_date()
    return render(request, 'Attendance_app/settings_app.html',
                  {'month': month, 'year': year, 'profile_users': profile_users})


def confirmation_vacation(request):
    staff = request.user

    users = CustomUser.objects.filter(created_who=staff)

    profile_not_vacation = Profile.objects.filter(user__in=users, vacation__check_by_employer=False)
    not_accepted_vacation = Vacation.objects.filter(user__in=users, check_by_employer=False)
    return render(request, 'Attendance_app/vacation_accept.html',
                  {'users': users, 'not_accepted_vacation': not_accepted_vacation,
                   'profile_not_vacation': profile_not_vacation})


def all_vacations(request):
    staff = request.user

    users = CustomUser.objects.filter(created_who=staff)
    vacation = Vacation.objects.filter(user__in=users)
    return render(request, 'Attendance_app/vacation_accept.html',
                  {'users': users, 'vacation': vacation})


def not_accepted_vacation(request, pk):
    vacation = Vacation.objects.get(id=pk)
    vacation.delete()
    return redirect(reverse('managing:check_vacation_confirmation'))


def accepted_vacation(request, pk):
    vacation = Vacation.objects.get(id=pk)
    vacation.check_by_employer = True
    vacation.save()
    return redirect(reverse('managing:check_vacation_confirmation'))
